import pymongo
import os
import pytz
import openpyxl
import traceback
import jieba
import urllib.request
from datetime import datetime
from constants import TaskStates, TaskTypes, enum_set, CUSTOMER_XLSX_HEADER_MAP
from config import Config

def process_task(task_id, max_retrial=5):
  mongo_client = pymongo.MongoClient(Config.MONGO_MAIN_URI)    
  db = mongo_client.get_database()
  task_col = db.schedulertasks
  result = None
  for i in range(max_retrial):
    task = task_col.find_one_and_update(
      { "_id": task_id },
      { 
        "$set": {
          "state": TaskStates.running,
          "trial": i + 1,
          "run_at": datetime.now(pytz.UTC),
          "system_pid": os.getpid(),
        }
      }
    )
    try:
      if task['task_type'] == TaskTypes.import_customer_xlsx:
        params = task['params']
        file_path = params['fs_path']
        result = process_member_xlsx(file_path, mongo_client=mongo_client)
        break
      else:
        raise ValueError('task_type should be one of %s' % enum_set(TaskTypes))
    except Exception as e:
      task_col.update_one(
        { "_id": task_id },
        { 
          "$set": {
            "state": TaskStates.failed,
            "message": traceback.format_exc(),
          }
        }
      )
      return
    
  # Mark task as completed
  task_col.update_one(
    { "_id": task_id },
    { 
      "$set": {
        "state": TaskStates.success,
        "message": result,
        "finished_at": datetime.now(pytz.UTC),
      }
    }
  )

def process_member_xlsx(file_path, mongo_client=None):
  zh_dict_path = './jieba.big'
  if not os.path.isfile(zh_dict_path):
    try:
      url = Config.LANGUAGE_DICT_URL
      urllib.request.urlretrieve(url, zh_dict_path)
    except:
      print ('failed to download dict from', Config.LANGUAGE_DICT_URL)
  if os.path.isfile(zh_dict_path):
    jieba.set_dictionary(zh_dict_path)
  
  if not mongo_client:
    mongo_client = pymongo.MongoClient(Config.MONGO_MAIN_URI)    

  db = mongo_client.get_database()
  member_col = db.customerinfos

  wb = openpyxl.load_workbook(file_path)
  ws = wb.active
  headers = [cell.value for cell in ws[1]]
  # Ensure all headers exist in the mapping
  column_map = {index: CUSTOMER_XLSX_HEADER_MAP.get(header, None) for index, header in enumerate(headers)}
  if None in column_map.values():
    missing_headers = [headers[i] for i, v in column_map.items() if v is None]
    raise ValueError("Warning: Unmapped headers found - %s" % (str(missing_headers), ))

  data_list = []
  for row in ws.iter_rows(min_row=2, values_only=True):
    data = {}
    text_fields_set = []

    for index, value in enumerate(row):
      field_name = column_map.get(index)
      if not field_name:
        continue
      # Convert Data Types
      if field_name == "gender":
        gender_code = value.upper().strip()
        if gender_code == 'M':
          data[field_name] = 1
        elif gender_code == 'F':
          data[field_name] = 2
        else:
          data[field_name] = None
      elif field_name == "civ_id":
        data[field_name] = value.upper()
      elif field_name == "birthday":
        data[field_name] = value
      elif field_name == "graduate_year":
        data[field_name] = int(value) if isinstance(value, (int, float)) else None
      elif field_name == "email":
        data[field_name] = str(value).lower() or ""
      else:
        data[field_name] = str(value).strip() if value else ""
        text_fields_set.append(data[field_name])

    tokens = list(jieba.cut(' '.join(t for t in text_fields_set if t.strip())))
    data['__tokens__'] = [t for t in tokens if t.strip()]
    data_list.append(data)

  # Insert into MongoDB
  if data_list:
    member_col.insert_many(data_list)
    return f"{len(data_list)} records processed"
